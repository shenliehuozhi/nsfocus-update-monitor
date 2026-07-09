"""Rule Diff API — /api/diff/* — live-fetches vendor portal HTML and computes
rule-package diff between two versions.

Three endpoints:
  GET  /api/diff/fetch?url=<vendor-url>[&force=1]
       Validate URL (must be from IPS/IDS/IDS-ICS listNewipsDetail path),
       check /tmp/rd_cache_<hash>.json for fresh (<24h) cache; if miss or
       ?force=1, live-fetch with PHPSESSID, parse all <table> blocks,
       persist to cache, return list of rule packages.

  POST /api/diff/compare
       Body: {packages: [{package_version, description_raw}, ...],
              from_version: str, to_version: str}
       Compute diff (parse_rules + diff_rules in src.diff).
       Pure compute; no DB.

  POST /api/diff/export
       Body: {diff_result: {...}, format: 'csv' | 'xlsx', filename: str?}
       Generate downloadable file in memory, return as Flask send_file.
       Stats sheet simplified to 4 rows: 新增/修改/删除/中间版本.
"""

import hashlib
import io
import json
import os
import time
from datetime import datetime
from typing import List, Optional

from flask import Blueprint, request, jsonify, g, send_file, Response

from src.web.auth import require_auth
from src.core.logger import get_logger

logger = get_logger('api.diff')

from src.collectors.nsfocus import (
    NsfocusCollector,
    SessionExpiredError,
    RedirectToLicenseError,
    BASE_URL,
)

from src.diff import (
    parse_rule_packages_from_html,
    parse_rules,
    parse_rules_waf,
    diff_rules,
    diff_to_csv_rows,
    compare_versions,
)

bp_diff = Blueprint('diff', __name__, url_prefix='/api/diff')


# === Constants ===
# Cache lives in /tmp until manually overwritten by ?force=1.
# No 24h TTL — the cache is the user's local copy of vendor portal data
# until they explicitly choose to re-fetch.
CACHE_DIR = '/tmp'
CACHE_PREFIX = 'rd_cache_'


# === Helpers ===

def _resolve_collect_cookie() -> tuple[bool, str | None, str | None]:
    try:
        from src.models.user_session import get_active_collect_sessions
        sessions = get_active_collect_sessions()
        if not sessions:
            return False, None, '无 collect 用途 Session, 请先在「Session 管理」中添加'
        sess = sessions.get('standard') or next(iter(sessions.values()))
        return True, sess['cookie_value'], None
    except Exception as e:
        logger.error(f'failed to resolve collect cookie: {e}')
        return False, None, f'读取 Session 失败: {e}'


def _normalize_vendor_url(url: str) -> str:
    url = (url or '').strip()
    if not url:
        return ''
    if url.startswith('http://') or url.startswith('https://'):
        return url
    if url.startswith('//'):
        return 'https:' + url
    if url.startswith('/'):
        return f'{BASE_URL}{url}'
    return f'{BASE_URL}/{url}'


def _cache_path_for(url: str) -> str:
    """Return the on-disk path for a given vendor URL's cache.

    Cache files are gzip-compressed to keep /tmp usage low (~10x smaller for
    rule pages with hundreds of description_raw blocks). We use a `.json.gz`
    suffix so it's obvious on disk.
    """
    h = hashlib.sha256(url.encode()).hexdigest()[:16]
    return os.path.join(CACHE_DIR, f'{CACHE_PREFIX}{h}.json.gz')


def _load_cache(url: str) -> Optional[dict]:
    """Return cached payload if a file exists for this URL.

    No TTL: cache persists until the user clicks "强制刷新" which overwrites it.
    Accepts both `.json.gz` (preferred) and legacy `.json` (back-compat).
    """
    import gzip
    p = _cache_path_for(url)
    if not os.path.exists(p):
        # Legacy uncompressed cache file?
        legacy = os.path.join(CACHE_DIR, f'{CACHE_PREFIX}{hashlib.sha256(url.encode()).hexdigest()[:16]}.json')
        if os.path.exists(legacy):
            p = legacy
            is_gz = False
        else:
            return None
    else:
        is_gz = True
    try:
        if is_gz:
            with gzip.open(p, 'rt', encoding='utf-8') as f:
                return json.load(f)
        else:
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.warning(f'cache load failed for {url}: {e}')
        return None


def _save_cache(url: str, payload: dict) -> None:
    """Persist a payload to disk as gzip-compressed JSON.

    Overwrites any existing cache file (and any legacy uncompressed file).
    Typical compression ratio on a rule-page dump is 10:1 — a 2.2 MB raw
    JSON shrinks to ~150 KB.
    """
    import gzip
    p = _cache_path_for(url)
    # Also remove any legacy uncompressed file for the same URL
    legacy = p[:-3] if p.endswith('.gz') else p  # strips .gz
    for candidate in (p, legacy):
        if candidate != p and os.path.exists(candidate):
            try:
                os.remove(candidate)
            except Exception:
                pass
    try:
        tmp = p + '.tmp'
        with gzip.open(tmp, 'wt', encoding='utf-8', compresslevel=6) as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, p)
    except Exception as e:
        logger.warning(f'cache save failed for {url}: {e}')


def _validate_diff_url(url: str) -> tuple[bool, str]:
    """Check that `url` is a known listNewipsDetail or listWafV...Detail path
    belonging to an IPS / IDS / WAF product in the live DB.

    Returns (ok, message):
      - (True, '')              URL is valid
      - (False, '具体错误...')  URL is not in the DB or not in scope
    Returns the source_kind ('ips' or 'waf') alongside ok when valid —
    callers use it to pick the right parser. (Returning 3 values via a dict
    would be cleaner; tuple keeps the existing signature contract.)
    """
    if not url.startswith('https://update.nsfocus.com/'):
        return False, '仅支持 update.nsfocus.com 域名', None

    path = url.replace('https://update.nsfocus.com', '').rstrip('/')

    # Decide which family of paths this is and the corresponding source whitelist
    if path.startswith('/update/listNewipsDetail/'):
        source_kind = 'ips'
        allowed_names = {
            '网络入侵防护系统(IPS)',
            '网络入侵检测系统(IDS)',
            '工控入侵检测系统(IDS-ICS)',
        }
    elif '/listWafV' in path and 'Detail' in path:
        source_kind = 'waf'
        allowed_names = {
            'WEB应用防护系统(WAF)',
            'WEB应用防护系统主机版(HWAF)',
        }
    else:
        return False, (
            f'当前工具仅支持 listNewipsDetail/ (IPS 规则) 或 listWafV...Detail '
            f'(WAF 规则) 路径,你给的是 {path}'
        ), None

    try:
        from src.models.database import query
        rows = query("SELECT id, name, package_type FROM content_sources WHERE is_active = 1")
    except Exception as e:
        logger.error(f'validate_diff_url: query content_sources failed: {e}')
        return False, f'无法读取产品配置: {e}', None

    matched_sources = []
    for r in rows:
        pt_raw = r['package_type']
        if not pt_raw:
            continue
        try:
            pt = json.loads(pt_raw)
        except Exception:
            continue
        if not isinstance(pt, dict):
            continue
        for p in (pt.get('paths') or []):
            if isinstance(p, dict) and (p.get('url') or '').rstrip('/') == path:
                matched_sources.append(r['name'] or '')
                break  # one match per source is enough

    if not matched_sources:
        return False, '该 URL 不在已知产品白名单中。请使用 IPS/IDS/WAF 产品的 vendor portal 链接', None

    hits = [s for s in matched_sources if s in allowed_names]
    if not hits:
        names = ', '.join(set(matched_sources))
        return False, f'该 URL 属于 {names}, 不在 IPS/IDS/WAF 产品范围内', None

    return True, '', source_kind


# === GET /api/diff/fetch ===

@bp_diff.route('/fetch', methods=['GET'])
@require_auth
def fetch_packages():
    """Live-fetch vendor portal HTML, parse all rule-package tables, return JSON.

    Query: ?url=https://update.nsfocus.com/update/listNewipsDetail/v/rule5.6.11_v2[&force=1]
    """
    url = request.args.get('url', '').strip()
    force = request.args.get('force', '').strip() in ('1', 'true', 'yes')

    if not url:
        return jsonify({'code': 400, 'message': '缺少 url 参数'}), 400

    full_url = _normalize_vendor_url(url)
    if not full_url.startswith('https://update.nsfocus.com/'):
        return jsonify({'code': 400, 'message': '仅支持 update.nsfocus.com 域名'}), 400

    ok, err, source_kind = _validate_diff_url(full_url)
    if not ok:
        return jsonify({'code': 400, 'message': err}), 400

    # 1) Try cache first (unless force=1)
    if not force:
        cached = _load_cache(full_url)
        if cached:
            cached['cache'] = 'hit'
            cached['source_kind'] = source_kind  # override in case cache predates the field
            logger.info(f'diff fetch cache HIT url={full_url}')
            return jsonify({'code': 0, 'data': cached})

    # 2) Live fetch
    ok, cookie, err = _resolve_collect_cookie()
    if not ok:
        return jsonify({'code': 400, 'message': err}), 400

    collector = NsfocusCollector()
    collector._set_cookie(cookie)

    try:
        resp = collector.session.get(full_url, timeout=30)
        resp.raise_for_status()
        if '/portal/' in resp.url or '/login' in resp.url.lower() or '登录' in resp.text[:200]:
            raise SessionExpiredError('Session invalid (login page redirect)')
        html = resp.text
    except SessionExpiredError as e:
        logger.warning(f'diff fetch session expired: {e}')
        return jsonify({'code': 401, 'message': 'PHPSESSID 已过期, 请在「Session 管理」中更新'}), 401
    except Exception as e:
        logger.error(f'diff fetch failed for {full_url}: {e}')
        return jsonify({'code': 500, 'message': f'拉取失败: {e}'}), 500

    try:
        packages = parse_rule_packages_from_html(html, full_url)
    except Exception as e:
        logger.error(f'diff parse failed: {e}')
        return jsonify({'code': 500, 'message': f'解析失败: {e}'}), 500

    def _sort_key(p):
        return (p.get('published_at', ''), p.get('package_version', ''))
    packages.sort(key=_sort_key)

    payload = {
        'url': full_url,
        'source_kind': source_kind,
        'fetched_at': datetime.utcnow().isoformat() + 'Z',
        'cache': 'miss',
        'packages': packages,
    }
    _save_cache(full_url, payload)
    logger.info(f'diff fetch OK url={full_url} kind={source_kind} packages={len(packages)} bytes={len(html)} cache_saved')
    return jsonify({'code': 0, 'data': payload})


# === POST /api/diff/compare ===

@bp_diff.route('/compare', methods=['POST'])
@require_auth
def compare_packages():
    """Compute diff between two rule packages, with optional intermediates.

    Body params:
      from_version, to_version: package versions to compare
      packages: list of {package_version, description_raw, ...}
      source_kind: 'ips' (default) or 'waf' — picks the rule-list parser
    """
    body = request.get_json(silent=True) or {}
    from_ver = (body.get('from_version') or '').strip()
    to_ver   = (body.get('to_version') or '').strip()
    packages = body.get('packages') or []
    source_kind = (body.get('source_kind') or 'ips').strip().lower()
    if source_kind not in ('ips', 'waf'):
        return jsonify({'code': 400, 'message': f'source_kind 必须是 ips 或 waf, 你给的是 {source_kind}'}), 400
    parser = parse_rules_waf if source_kind == 'waf' else parse_rules

    if not from_ver or not to_ver:
        return jsonify({'code': 400, 'message': 'from_version / to_version 必填'}), 400
    if compare_versions(from_ver, to_ver) == 0:
        return jsonify({'code': 400, 'message': 'from_version 和 to_version 不能相同'}), 400
    if not isinstance(packages, list) or not packages:
        return jsonify({'code': 400, 'message': 'packages 不能为空'}), 400

    by_ver = {}
    for p in packages:
        ver = (p.get('package_version') or '').strip()
        if not ver or not p.get('description_raw'):
            continue
        by_ver[ver] = {
            'package_version': ver,
            'description_raw': p['description_raw'],
            'published_at':    p.get('published_at', ''),
        }

    if from_ver not in by_ver:
        return jsonify({'code': 400, 'message': f'packages 中找不到 from_version={from_ver}'}), 400
    if to_ver not in by_ver:
        return jsonify({'code': 400, 'message': f'packages 中找不到 to_version={to_ver}'}), 400

    if compare_versions(from_ver, to_ver) > 0:
        from_ver, to_ver = to_ver, from_ver

    from_p = by_ver[from_ver]
    to_p   = by_ver[to_ver]

    intermediates = []
    for ver, p in by_ver.items():
        if ver in (from_ver, to_ver):
            continue
        if compare_versions(from_ver, ver) == -1 and compare_versions(ver, to_ver) == -1:
            intermediates.append((ver, parser(p['description_raw'])))
    intermediates.sort(key=lambda x: [int(s) if s.isdigit() else s for s in x[0].split('.')])

    from_parsed = parser(from_p['description_raw'])
    to_parsed   = parser(to_p['description_raw'])
    result = diff_rules(from_ver, to_ver, from_parsed, to_parsed, intermediates)

    result['stats']['from_has_rule_blocks'] = bool(from_parsed.get('added') or from_parsed.get('updated'))
    result['stats']['to_has_rule_blocks']   = bool(to_parsed.get('added')   or to_parsed.get('updated'))
    result['stats']['source_kind']          = source_kind

    logger.info(
        f'diff compare kind={source_kind} {from_ver} -> {to_ver} '
        f'intermediates={len(intermediates)} '
        f'new={len(result["pure_new"])} updated={len(result["pure_updated"])} '
        f'deleted={len(result["pure_deleted"])}'
    )
    return jsonify({'code': 0, 'data': result})


# === POST /api/diff/export ===

@bp_diff.route('/export', methods=['POST'])
@require_auth
def export_diff():
    """Generate a downloadable file (CSV or xlsx) for a diff result."""
    body = request.get_json(silent=True) or {}
    diff_result = body.get('diff_result') or {}
    fmt = (body.get('format') or 'csv').lower()
    filename = (body.get('filename') or 'rule_diff').strip()

    if not diff_result:
        return jsonify({'code': 400, 'message': 'diff_result 不能为空'}), 400

    rows = diff_to_csv_rows(diff_result)

    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({'code': 500, 'message': 'openpyxl 未安装, 请 pip install openpyxl'}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = 'Rule Diff'
        headers = ['类型', '规则号', '规则名称', '首次新增于', '更新于']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
            cell.alignment = Alignment(horizontal='center')

        type_color = {'新增': 'C8E6C9', '更新': 'BBDEFB', '删除': 'FFCDD2'}
        for row in rows:
            ws.append([row.get(h, '') or '' for h in headers])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor=type_color.get(row.get('类型', ''), 'FFFFFF'))

        widths = [8, 12, 60, 18, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

        # Stats sheet — simplified to 4 rows: 新增 / 修改 / 删除 / 中间版本
        ws2 = wb.create_sheet('Stats')
        stats = diff_result.get('stats', {})
        intermediates = stats.get('intermediate_versions', []) or []
        ws2.append(['指标', '值'])
        ws2.append(['新增规则数', len(diff_result.get('pure_new', []))])
        ws2.append(['修改规则数', len(diff_result.get('pure_updated', []))])
        ws2.append(['删除规则数', len(diff_result.get('pure_deleted', []))])
        ws2.append(['中间版本数', len(intermediates)])
        if intermediates:
            ws2.append(['中间版本列表', ', '.join(intermediates)])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
            cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions['A'].width = 24
        ws2.column_dimensions['B'].width = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = ''.join(c for c in filename if c.isalnum() or c in '._-') or 'rule_diff'
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'{safe_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    # CSV default — with BOM so Excel on Windows opens Chinese chars correctly
    import csv
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=['类型', '规则号', '规则名称', '首次新增于', '更新于'])
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    csv_text = '\ufeff' + buf.getvalue()

    safe_name = ''.join(c for c in filename if c.isalnum() or c in '._-') or 'rule_diff'
    return Response(
        csv_text,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{safe_name}.csv"'},
    )


# === GET /api/diff/cache-status — diagnostic endpoint ===
#   Returns whether a cache entry exists for the given URL (no auth required
#   for diagnostic; but still require_auth for consistency).
@bp_diff.route('/cache-status', methods=['GET'])
@require_auth
def cache_status():
    """Diagnostic: report whether a cache file exists for the given URL.

    No TTL: cache persists until the user clicks "强制刷新" which overwrites it.
    """
    url = request.args.get('url', '').strip()
    if not url:
        return jsonify({'code': 400, 'message': '缺少 url 参数'}), 400
    full_url = _normalize_vendor_url(url)
    p = _cache_path_for(full_url)
    if not os.path.exists(p):
        return jsonify({'code': 0, 'data': {'exists': False}})
    mtime = os.path.getmtime(p)
    age = int(time.time() - mtime)
    return jsonify({
        'code': 0,
        'data': {
            'exists': True,
            'path': p,
            'mtime': datetime.utcfromtimestamp(mtime).isoformat() + 'Z',
            'age_seconds': age,
        },
    })